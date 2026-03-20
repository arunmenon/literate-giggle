"""Class management, enrollment, and exam assignment routes."""

import csv
import io
import math
import re
import secrets
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from ...core.config import settings
from ...core.database import get_db
from ...core.rate_limit import limiter
from ...models.user import User, StudentProfile, UserRole
from ...models.exam import QuestionPaper
from ...models.workspace import (
    ClassGroup, Enrollment, ExamAssignment,
    Workspace, WorkspaceMember, generate_invite_code,
)
from ...schemas.workspace import (
    ClassGroupCreate, ClassGroupResponse,
    EnrollStudentRequest, EnrollmentResponse,
    ExamAssignmentCreate, ExamAssignmentResponse,
    JoinClassRequest, JoinClassResponse,
    CopyRosterRequest, CopyRosterResponse,
    BulkEnrollRequest, BulkEnrollResponse,
    ImportResponse,
    InviteLinkRequest, InviteLinkResponse,
    PaginatedEnrollmentResponse,
)
from ..deps import get_current_user, require_teacher_or_admin, get_current_workspace

router = APIRouter(prefix="/classes", tags=["Class Management"])


@router.post("", response_model=ClassGroupResponse, status_code=201)
async def create_class(
    data: ClassGroupCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Create a class in the active workspace."""
    ws_id = get_current_workspace(user)
    if not ws_id:
        raise HTTPException(status_code=400, detail="No active workspace")

    class_group = ClassGroup(
        workspace_id=ws_id,
        name=data.name,
        grade=data.grade,
        section=data.section,
        subject=data.subject,
        academic_year=data.academic_year,
        teacher_id=user.id,
    )
    db.add(class_group)
    await db.flush()

    return ClassGroupResponse(
        id=class_group.id,
        workspace_id=class_group.workspace_id,
        name=class_group.name,
        grade=class_group.grade,
        section=class_group.section,
        subject=class_group.subject,
        academic_year=class_group.academic_year,
        teacher_id=class_group.teacher_id,
        join_code=class_group.join_code,
        join_code_active=class_group.join_code_active,
        student_count=0,
        created_at=class_group.created_at,
    )


@router.post("/join", response_model=JoinClassResponse, status_code=200)
@limiter.limit("10/minute")
async def join_class(
    request: Request,
    data: JoinClassRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Student self-join a class by join code. Auto-joins the parent workspace."""
    # Find class by join_code
    result = await db.execute(
        select(ClassGroup).where(
            ClassGroup.join_code == data.join_code,
            ClassGroup.join_code_active == True,
        )
    )
    class_group = result.scalar_one_or_none()
    if not class_group:
        raise HTTPException(status_code=404, detail="Invalid or inactive join code")

    # Check max_students cap
    if class_group.max_students is not None:
        count_result = await db.execute(
            select(func.count(Enrollment.id)).where(
                Enrollment.class_id == class_group.id,
                Enrollment.is_active == True,
            )
        )
        current_count = count_result.scalar() or 0
        if current_count >= class_group.max_students:
            raise HTTPException(
                status_code=409,
                detail=f"Class is full (max {class_group.max_students} students)",
            )

    # Get student profile
    profile_result = await db.execute(
        select(StudentProfile).where(StudentProfile.user_id == user.id)
    )
    profile = profile_result.scalar_one_or_none()
    if not profile:
        raise HTTPException(status_code=400, detail="Only students can join classes")

    # Check not already enrolled
    existing = await db.execute(
        select(Enrollment).where(
            Enrollment.class_id == class_group.id,
            Enrollment.student_id == profile.id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Already enrolled in this class")

    # Create enrollment
    enrollment = Enrollment(class_id=class_group.id, student_id=profile.id)
    db.add(enrollment)

    # Auto-join workspace if not already a member
    ws_member_result = await db.execute(
        select(WorkspaceMember).where(
            WorkspaceMember.workspace_id == class_group.workspace_id,
            WorkspaceMember.user_id == user.id,
        )
    )
    if not ws_member_result.scalar_one_or_none():
        db.add(WorkspaceMember(
            workspace_id=class_group.workspace_id,
            user_id=user.id,
            role="student",
        ))

    # Set active workspace
    user.active_workspace_id = class_group.workspace_id
    await db.flush()

    # Fetch workspace name for response
    ws = await db.get(Workspace, class_group.workspace_id)

    return JoinClassResponse(
        class_id=class_group.id,
        class_name=class_group.name,
        workspace_id=class_group.workspace_id,
        workspace_name=ws.name if ws else "",
        message=f"Successfully joined {class_group.name}",
    )


@router.post("/{class_id}/regenerate-code", response_model=ClassGroupResponse)
async def regenerate_join_code(
    class_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Generate a new join code for a class (teacher only)."""
    class_group = await db.get(ClassGroup, class_id)
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    class_group.join_code = generate_invite_code()
    await db.flush()

    # Get student count
    count_result = await db.execute(
        select(func.count(Enrollment.id)).where(
            Enrollment.class_id == class_id,
            Enrollment.is_active == True,
        )
    )
    student_count = count_result.scalar() or 0

    return ClassGroupResponse(
        id=class_group.id,
        workspace_id=class_group.workspace_id,
        name=class_group.name,
        grade=class_group.grade,
        section=class_group.section,
        subject=class_group.subject,
        academic_year=class_group.academic_year,
        teacher_id=class_group.teacher_id,
        join_code=class_group.join_code,
        join_code_active=class_group.join_code_active,
        student_count=student_count,
        created_at=class_group.created_at,
    )


@router.patch("/{class_id}/join-code")
async def toggle_join_code(
    class_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Toggle join_code_active for a class (teacher only)."""
    class_group = await db.get(ClassGroup, class_id)
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    class_group.join_code_active = not class_group.join_code_active
    await db.flush()

    return {
        "class_id": class_group.id,
        "join_code": class_group.join_code,
        "join_code_active": class_group.join_code_active,
    }


@router.get("/{class_id}/qr-code")
async def get_qr_code(
    class_id: int,
    size: int = Query(default=300, ge=100, le=1000),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Generate a QR code image for the class join link."""
    import qrcode

    class_group = await db.get(ClassGroup, class_id)
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    join_url = f"{settings.FRONTEND_URL}/join/{class_group.join_code}"

    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(join_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    # Resize to requested size
    img = img.resize((size, size))

    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)

    return StreamingResponse(buffer, media_type="image/png")


@router.post("/{class_id}/copy-roster", response_model=CopyRosterResponse)
async def copy_roster(
    class_id: int,
    data: CopyRosterRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Copy active enrollments from a source class. Both classes must be in the same workspace."""
    target_class = await db.get(ClassGroup, class_id)
    if not target_class:
        raise HTTPException(status_code=404, detail="Target class not found")

    source_class = await db.get(ClassGroup, data.source_class_id)
    if not source_class:
        raise HTTPException(status_code=404, detail="Source class not found")

    if target_class.workspace_id != source_class.workspace_id:
        raise HTTPException(status_code=400, detail="Both classes must be in the same workspace")

    # Get active enrollments from source
    source_enrollments = await db.execute(
        select(Enrollment).where(
            Enrollment.class_id == data.source_class_id,
            Enrollment.is_active == True,
        )
    )
    source_students = source_enrollments.scalars().all()

    # Get existing enrollments in target (any status) to skip duplicates
    existing_result = await db.execute(
        select(Enrollment.student_id).where(Enrollment.class_id == class_id)
    )
    existing_student_ids = {row[0] for row in existing_result.all()}

    copied = 0
    skipped = 0
    for enrollment in source_students:
        if enrollment.student_id in existing_student_ids:
            skipped += 1
        else:
            db.add(Enrollment(class_id=class_id, student_id=enrollment.student_id))
            copied += 1

    await db.flush()

    return CopyRosterResponse(copied=copied, skipped=skipped, total=copied + skipped)


async def _bulk_enroll_logic(
    class_group: ClassGroup,
    emails: list[str],
    db: AsyncSession,
) -> BulkEnrollResponse:
    """Shared bulk enrollment logic used by both bulk and import endpoints.

    Looks up users by email in a single IN query, enrolls those with
    StudentProfile, auto-joins workspace, and categorises results.
    """
    enrolled: list[str] = []
    already_enrolled: list[str] = []
    not_found: list[str] = []
    errors: list[str] = []

    # Deduplicate and normalise
    unique_emails = list({e.strip().lower() for e in emails if e and e.strip()})
    if not unique_emails:
        return BulkEnrollResponse(
            enrolled=enrolled, already_enrolled=already_enrolled,
            not_found=not_found, errors=errors,
        )

    # Batch lookup users by email (avoids N+1)
    user_result = await db.execute(
        select(User).where(User.email.in_(unique_emails))
    )
    users_by_email: dict[str, User] = {u.email.lower(): u for u in user_result.scalars().all()}

    # Batch lookup student profiles for found users
    found_user_ids = [u.id for u in users_by_email.values()]
    profile_result = await db.execute(
        select(StudentProfile).where(StudentProfile.user_id.in_(found_user_ids))
    ) if found_user_ids else None
    profiles_by_user_id: dict[int, StudentProfile] = {}
    if profile_result:
        profiles_by_user_id = {p.user_id: p for p in profile_result.scalars().all()}

    # Batch lookup existing enrollments for this class
    profile_ids = [p.id for p in profiles_by_user_id.values()]
    existing_enrollment_profile_ids: set[int] = set()
    if profile_ids:
        existing_result = await db.execute(
            select(Enrollment.student_id).where(
                Enrollment.class_id == class_group.id,
                Enrollment.student_id.in_(profile_ids),
            )
        )
        existing_enrollment_profile_ids = {row[0] for row in existing_result.all()}

    # Batch lookup existing workspace members
    existing_ws_member_user_ids: set[int] = set()
    if found_user_ids:
        ws_member_result = await db.execute(
            select(WorkspaceMember.user_id).where(
                WorkspaceMember.workspace_id == class_group.workspace_id,
                WorkspaceMember.user_id.in_(found_user_ids),
            )
        )
        existing_ws_member_user_ids = {row[0] for row in ws_member_result.all()}

    # Process each email
    for email in unique_emails:
        user = users_by_email.get(email)
        if not user:
            not_found.append(email)
            continue

        profile = profiles_by_user_id.get(user.id)
        if not profile:
            errors.append(f"{email}: user is not a student")
            continue

        if profile.id in existing_enrollment_profile_ids:
            already_enrolled.append(email)
            continue

        # Create enrollment
        db.add(Enrollment(class_id=class_group.id, student_id=profile.id))
        existing_enrollment_profile_ids.add(profile.id)

        # Auto-join workspace if not already a member
        if user.id not in existing_ws_member_user_ids:
            db.add(WorkspaceMember(
                workspace_id=class_group.workspace_id,
                user_id=user.id,
                role="student",
            ))
            existing_ws_member_user_ids.add(user.id)

        enrolled.append(email)

    await db.flush()

    return BulkEnrollResponse(
        enrolled=enrolled,
        already_enrolled=already_enrolled,
        not_found=not_found,
        errors=errors,
    )


# ── Task 2.1: Bulk Enrollment ──


@router.post("/{class_id}/students/bulk", response_model=BulkEnrollResponse)
async def bulk_enroll_students(
    class_id: int,
    data: BulkEnrollRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Batch enroll students by email list. Uses IN queries (no N+1)."""
    class_group = await db.get(ClassGroup, class_id)
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    # Validate email format
    email_pattern = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    invalid_emails = [e for e in data.emails if not email_pattern.match(e.strip())]
    if invalid_emails:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid email format: {', '.join(invalid_emails[:5])}",
        )

    return await _bulk_enroll_logic(class_group, data.emails, db)


# ── Task 2.2: CSV/Excel Import ──

MAX_IMPORT_ROWS = 500
MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024  # 5 MB


@router.post("/{class_id}/students/import", response_model=ImportResponse)
async def import_students(
    class_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Import students from a CSV or XLSX file. Required columns: name, email."""
    class_group = await db.get(ClassGroup, class_id)
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    # Read file content with size limit
    content = await file.read()
    if len(content) > MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File exceeds 5 MB limit")

    filename = (file.filename or "").lower()
    is_xlsx = filename.endswith(".xlsx") or (
        file.content_type and "spreadsheet" in file.content_type
    )

    rows: list[dict[str, str]] = []
    parse_errors: list[str] = []

    if is_xlsx:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status_code=400, detail="Empty workbook")

            # Read headers from first row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip().lower() if h else "" for h in header_row]

            if "email" not in headers:
                raise HTTPException(
                    status_code=400,
                    detail="Missing required column: email. Found columns: " + ", ".join(headers),
                )

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row_idx - 1 > MAX_IMPORT_ROWS:
                    parse_errors.append(f"Stopped at row {MAX_IMPORT_ROWS}: max rows exceeded")
                    break
                row_dict = {headers[i]: str(cell).strip() if cell else "" for i, cell in enumerate(row) if i < len(headers)}
                if row_dict.get("email"):
                    rows.append(row_dict)
            wb.close()
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {exc}")
    else:
        # CSV parsing
        try:
            text = content.decode("utf-8-sig")  # handle BOM
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise HTTPException(status_code=400, detail="Empty CSV file")

            normalised_fields = [f.strip().lower() for f in reader.fieldnames]
            if "email" not in normalised_fields:
                raise HTTPException(
                    status_code=400,
                    detail="Missing required column: email. Found columns: " + ", ".join(normalised_fields),
                )

            # Re-map field names to lowercase
            for row_idx, row in enumerate(reader, start=2):
                if row_idx - 1 > MAX_IMPORT_ROWS:
                    parse_errors.append(f"Stopped at row {MAX_IMPORT_ROWS}: max rows exceeded")
                    break
                normalised = {k.strip().lower(): (v.strip() if v else "") for k, v in row.items()}
                if normalised.get("email"):
                    rows.append(normalised)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {exc}")

    if not rows:
        return ImportResponse(total_rows=0, enrolled=0, skipped=0, errors=parse_errors)

    emails = [row["email"] for row in rows]
    result = await _bulk_enroll_logic(class_group, emails, db)

    return ImportResponse(
        total_rows=len(rows),
        enrolled=len(result.enrolled),
        skipped=len(result.already_enrolled) + len(result.not_found),
        errors=parse_errors + result.errors + [f"{e}: not registered" for e in result.not_found],
    )


# ── Task 2.3: Invite Link ──


@router.post("/{class_id}/invite-link", response_model=InviteLinkResponse)
async def create_invite_link(
    class_id: int,
    data: InviteLinkRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Generate a shareable invite link with expiry for a class."""
    class_group = await db.get(ClassGroup, class_id)
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    token = secrets.token_urlsafe(48)  # 64-char URL-safe token
    expires_at = datetime.now(timezone.utc) + timedelta(hours=data.expires_in_hours)

    class_group.invite_link_token = token
    class_group.link_expires_at = expires_at
    await db.flush()

    invite_url = f"{settings.FRONTEND_URL}/join/link/{token}"

    return InviteLinkResponse(invite_url=invite_url, expires_at=expires_at)


@router.post("/join/link/{token}", response_model=JoinClassResponse, status_code=200)
@limiter.limit("10/minute")
async def join_via_invite_link(
    request: Request,
    token: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Join a class via shareable invite link. Expired links return 410 Gone."""
    result = await db.execute(
        select(ClassGroup).where(ClassGroup.invite_link_token == token)
    )
    class_group = result.scalar_one_or_none()
    if not class_group:
        raise HTTPException(status_code=404, detail="Invalid invite link")

    # Check expiry
    if class_group.link_expires_at is None or class_group.link_expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=410, detail="Invite link has expired")

    # Get student profile
    profile_result = await db.execute(
        select(StudentProfile).where(StudentProfile.user_id == user.id)
    )
    profile = profile_result.scalar_one_or_none()
    if not profile:
        raise HTTPException(status_code=400, detail="Only students can join classes")

    # Check not already enrolled
    existing = await db.execute(
        select(Enrollment).where(
            Enrollment.class_id == class_group.id,
            Enrollment.student_id == profile.id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Already enrolled in this class")

    # Check max_students cap
    if class_group.max_students is not None:
        count_result = await db.execute(
            select(func.count(Enrollment.id)).where(
                Enrollment.class_id == class_group.id,
                Enrollment.is_active == True,
            )
        )
        current_count = count_result.scalar() or 0
        if current_count >= class_group.max_students:
            raise HTTPException(
                status_code=409,
                detail=f"Class is full (max {class_group.max_students} students)",
            )

    # Create enrollment
    db.add(Enrollment(class_id=class_group.id, student_id=profile.id))

    # Auto-join workspace if not already a member
    ws_member_result = await db.execute(
        select(WorkspaceMember).where(
            WorkspaceMember.workspace_id == class_group.workspace_id,
            WorkspaceMember.user_id == user.id,
        )
    )
    if not ws_member_result.scalar_one_or_none():
        db.add(WorkspaceMember(
            workspace_id=class_group.workspace_id,
            user_id=user.id,
            role="student",
        ))

    user.active_workspace_id = class_group.workspace_id
    await db.flush()

    ws = await db.get(Workspace, class_group.workspace_id)

    return JoinClassResponse(
        class_id=class_group.id,
        class_name=class_group.name,
        workspace_id=class_group.workspace_id,
        workspace_name=ws.name if ws else "",
        message=f"Successfully joined {class_group.name}",
    )


@router.get("", response_model=list[ClassGroupResponse])
async def list_classes(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """List classes in the active workspace."""
    ws_id = get_current_workspace(user)
    if not ws_id:
        return []

    query = select(ClassGroup).where(ClassGroup.workspace_id == ws_id)
    # Teachers see only their own classes unless they're admin/owner
    ws_role = getattr(user, "_workspace_role", None)
    if user.role == UserRole.TEACHER and ws_role not in ("owner", "admin"):
        query = query.where(ClassGroup.teacher_id == user.id)

    # Single aggregated query: join classes with enrollment counts to avoid N+1
    count_subq = (
        select(
            Enrollment.class_id,
            func.count(Enrollment.id).label("student_count"),
        )
        .where(Enrollment.is_active == True)
        .group_by(Enrollment.class_id)
        .subquery()
    )
    combined = (
        query.outerjoin(count_subq, ClassGroup.id == count_subq.c.class_id)
        .add_columns(func.coalesce(count_subq.c.student_count, 0).label("student_count"))
        .order_by(ClassGroup.created_at.desc())
    )
    result = await db.execute(combined)
    rows = result.all()

    return [
        ClassGroupResponse(
            id=cls.id,
            workspace_id=cls.workspace_id,
            name=cls.name,
            grade=cls.grade,
            section=cls.section,
            subject=cls.subject,
            academic_year=cls.academic_year,
            teacher_id=cls.teacher_id,
            join_code=cls.join_code,
            join_code_active=cls.join_code_active,
            student_count=count,
            created_at=cls.created_at,
        )
        for cls, count in rows
    ]


@router.get("/{class_id}/students", response_model=PaginatedEnrollmentResponse)
async def list_class_students(
    class_id: int,
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=25, ge=1, le=100),
    search: Optional[str] = Query(default=None),
    sort_by: str = Query(default="name"),
    sort_order: str = Query(default="asc"),
    active_only: bool = Query(default=True),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """List enrolled students with pagination, search, and sorting."""
    # Base query with joins
    base_query = (
        select(Enrollment, StudentProfile, User)
        .join(StudentProfile, Enrollment.student_id == StudentProfile.id)
        .join(User, StudentProfile.user_id == User.id)
        .where(Enrollment.class_id == class_id)
    )

    if active_only:
        base_query = base_query.where(Enrollment.is_active == True)

    # Search filter (case-insensitive on name or email)
    if search:
        search_term = f"%{search}%"
        base_query = base_query.where(
            or_(
                User.full_name.ilike(search_term),
                User.email.ilike(search_term),
            )
        )

    # Count total before pagination
    count_query = select(func.count()).select_from(base_query.subquery())
    total = (await db.execute(count_query)).scalar() or 0

    # Sorting
    sort_column_map = {
        "name": User.full_name,
        "email": User.email,
        "enrolled_at": Enrollment.enrolled_at,
    }
    sort_col = sort_column_map.get(sort_by, User.full_name)
    if sort_order == "desc":
        base_query = base_query.order_by(sort_col.desc())
    else:
        base_query = base_query.order_by(sort_col.asc())

    # Pagination
    offset = (page - 1) * per_page
    paginated_query = base_query.offset(offset).limit(per_page)

    result = await db.execute(paginated_query)
    rows = result.all()

    students = [
        EnrollmentResponse(
            id=enrollment.id,
            class_id=enrollment.class_id,
            student_id=enrollment.student_id,
            student_name=u.full_name,
            student_email=u.email,
            enrolled_at=enrollment.enrolled_at,
            is_active=enrollment.is_active,
        )
        for enrollment, profile, u in rows
    ]

    total_pages = math.ceil(total / per_page) if total > 0 else 1

    return PaginatedEnrollmentResponse(
        students=students,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
    )


@router.post("/{class_id}/enroll", response_model=EnrollmentResponse, status_code=201)
async def enroll_student(
    class_id: int,
    data: EnrollStudentRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Enroll a student in a class by user_id or email."""
    class_group = await db.get(ClassGroup, class_id)
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    # Find student by user_id or email
    if data.student_user_id:
        student_user_result = await db.execute(
            select(User).where(User.id == data.student_user_id)
        )
        student_user = student_user_result.scalar_one_or_none()
    elif data.email:
        student_user_result = await db.execute(
            select(User).where(User.email == data.email)
        )
        student_user = student_user_result.scalar_one_or_none()
    else:
        raise HTTPException(status_code=400, detail="Provide student_user_id or email")

    if not student_user:
        raise HTTPException(status_code=404, detail="Student user not found")

    profile_result = await db.execute(
        select(StudentProfile).where(StudentProfile.user_id == student_user.id)
    )
    profile = profile_result.scalar_one_or_none()
    if not profile:
        raise HTTPException(status_code=400, detail="User is not a student")

    # Check if already enrolled
    existing = await db.execute(
        select(Enrollment).where(
            Enrollment.class_id == class_id,
            Enrollment.student_id == profile.id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Student already enrolled")

    enrollment = Enrollment(
        class_id=class_id,
        student_id=profile.id,
    )
    db.add(enrollment)
    await db.flush()

    return EnrollmentResponse(
        id=enrollment.id,
        class_id=enrollment.class_id,
        student_id=enrollment.student_id,
        student_name=student_user.full_name,
        student_email=student_user.email,
        enrolled_at=enrollment.enrolled_at,
        is_active=enrollment.is_active,
    )


@router.delete("/{class_id}/enroll/{student_id}", status_code=204)
async def remove_student(
    class_id: int,
    student_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Remove a student from a class (soft delete)."""
    result = await db.execute(
        select(Enrollment).where(
            Enrollment.class_id == class_id,
            Enrollment.student_id == student_id,
        )
    )
    enrollment = result.scalar_one_or_none()
    if not enrollment:
        raise HTTPException(status_code=404, detail="Enrollment not found")
    enrollment.is_active = False
    await db.flush()


@router.post("/{class_id}/assign-exam", response_model=ExamAssignmentResponse, status_code=201)
async def assign_exam(
    class_id: int,
    data: ExamAssignmentCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_teacher_or_admin),
):
    """Assign a paper to a class with optional schedule."""
    class_group = await db.get(ClassGroup, class_id)
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    paper = await db.get(QuestionPaper, data.paper_id)
    if not paper:
        raise HTTPException(status_code=404, detail="Paper not found")

    assignment = ExamAssignment(
        paper_id=data.paper_id,
        class_id=class_id,
        assigned_by=user.id,
        status=data.status,
        label=data.label,
        start_at=data.start_at,
        end_at=data.end_at,
        is_practice=data.is_practice,
    )
    db.add(assignment)
    await db.flush()

    return ExamAssignmentResponse(
        id=assignment.id,
        paper_id=assignment.paper_id,
        paper_title=paper.title,
        class_id=assignment.class_id,
        class_name=class_group.name,
        assigned_by=assignment.assigned_by,
        status=assignment.status,
        label=assignment.label,
        start_at=assignment.start_at,
        end_at=assignment.end_at,
        is_practice=assignment.is_practice,
        created_at=assignment.created_at,
    )


@router.get("/{class_id}/assignments", response_model=list[ExamAssignmentResponse])
async def list_assignments(
    class_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """List exam assignments for a class."""
    result = await db.execute(
        select(ExamAssignment, QuestionPaper, ClassGroup)
        .join(QuestionPaper, ExamAssignment.paper_id == QuestionPaper.id)
        .join(ClassGroup, ExamAssignment.class_id == ClassGroup.id)
        .where(ExamAssignment.class_id == class_id)
        .order_by(ExamAssignment.created_at.desc())
    )
    rows = result.all()
    return [
        ExamAssignmentResponse(
            id=assignment.id,
            paper_id=assignment.paper_id,
            paper_title=paper.title,
            class_id=assignment.class_id,
            class_name=cls.name,
            assigned_by=assignment.assigned_by,
            status=assignment.status,
            label=assignment.label,
            start_at=assignment.start_at,
            end_at=assignment.end_at,
            is_practice=assignment.is_practice,
            created_at=assignment.created_at,
        )
        for assignment, paper, cls in rows
    ]
